import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def setup_excel(file_path):
    """Create Excel if not exists and ensure required sheets."""
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        default_sheet.title = "Alerts"
        default_sheet.append([
            "Group Name", "Alert Name", "Ignore Duplicates Alerts within minutes",
            "Threshold /Value (nmi)", "Scheduled Days", "Start Time", "End Time",
            "Delivery action email/sms"
        ])
        wb.save(file_path)

    required_sheets = {
        "MainDeviceList": ["Status", "Last Message Date", "Client Name", "Device Name", "IMEI", "SMS Number",
                           "Last Reported Date", "Modified By", "Modified", "VIN", "Group Name", "Created Date",
                           "Hardware Name", "IP Address"],
        "Alerts": ["Group Name", "Alert Name", "Ignore Duplicates Alerts within minutes",
                   "Threshold /Value (nmi)", "Scheduled Days", "Start Time", "End Time",
                   "Delivery action email/sms"],
        "Reports": ["Report Name", "Email to", "Email subject", "Report Language",
                    "Report File format", "Send time", "Repeat Frequency", "Start Date"],
        "Maintenance": ["Operation", "Description"],
        "Credentials": ["Full Name", "Username", "Associated Group Name",
                        "Group Filter", "Associate Items"],
        "Results": ["Timestamp", "Module Name", "Status", "Message"]
    }

    for sheet_name, headers in required_sheets.items():
        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(sheet_name)
            sheet.append(headers)
        else:
            sheet = wb[sheet_name]
            if sheet.max_row == 0:
                sheet.append(headers)

    wb.save(file_path)
    print(f"✅ Excel ready: {file_path}")
    return wb  # Return workbook object for further use

def append_rows_to_sheet(file_path, sheet_name, rows):
    """Append multiple rows to an Excel sheet."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"{file_path} not found. Call setup_excel() first.")
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' does not exist.")
    sheet = wb[sheet_name]
    for row in rows:
        # Convert all elements to string to avoid WebElement issues
        row_str = [str(cell) for cell in row]
        sheet.append(row_str)
    wb.save(file_path)
    print(f"✅ Appended {len(rows)} rows to '{sheet_name}'")
