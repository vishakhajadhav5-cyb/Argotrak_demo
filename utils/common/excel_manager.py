from openpyxl import Workbook, load_workbook
import os

class ExcelManager:
    def __init__(self, file_path):
        self.file_path = file_path

        if os.path.exists(self.file_path):
            self.wb = load_workbook(self.file_path)
        else:
            self.wb = Workbook()
            # Ensure at least Alerts sheet exists
            default_sheet = self.wb.active
            default_sheet.title = "Alerts"
            default_sheet.append([
                "Group Name", "Alert Name", "Ignore Duplicates Alerts within minutes",
                "Threshold /Value (nmi)", "Scheduled Days", "Start Time", "End Time",
                "Delivery action email/sms"
            ])
            self.wb.save(self.file_path)

    def ensure_sheets(self):
        required_sheets = {
        "MainDeviceList": [
            "Status", "Last Message Date", "Client Name", "Device Name", "IMEI", "SMS Number",
            "Last Reported Date", "Modified By", "Modified", "VIN", "Group Name",
            "Created Date", "Hardware Name", "IP Address"
        ],

        "Alerts": [
            "Group Name", "Alert Name", "Ignore Duplicates Alerts within minutes",
            "Threshold /Value (nmi)", "Scheduled Days", "Start Time", "End Time",
            "Delivery action email/sms"
        ],

        "Reports": [
            "Report Name", "Email to", "Email subject", "Report Language",
            "Report File format", "Send time", "Repeat Frequency", "Start Date"
        ],

        "Maintenance": [
            "Operation", "Description"
        ],

        "Credentials": [
            "Full Name", "Username", "Associated Group Name",
            "Group Filter", "Associate Items"
        ],

        "Results": [
            "Timestamp", "Module Name", "Status", "Message"
        ]
    }

        for sheet_name, headers in required_sheets.items():
            if sheet_name not in self.wb.sheetnames:
                sheet = self.wb.create_sheet(sheet_name)
                sheet.append(headers)
            else:
                sheet = self.wb[sheet_name]
                if sheet.max_row == 0:
                    sheet.append(headers)
        self.wb.save(self.file_path)

    def append_to_sheet(self, sheet_name, row_data):
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' does not exist. Call ensure_sheets() first.")
        sheet = self.wb[sheet_name]
        sheet.append(row_data)
        self.wb.save(self.file_path)
